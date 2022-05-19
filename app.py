from flask import Flask,jsonify,redirect,render_template,request,Response,session,url_for
from flask_cors import CORS, cross_origin   # pip install -U flask-cors
from flask_session import Session
import json
import logging
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.writer.excel import save_virtual_workbook
import pandas as pd

# https://www.imdb.com/interfaces/
# https://medium.com/analytics-vidhya/exploratory-data-analysis-imdb-dataset-cff0c3991ad5

# https://roytuts.com/autocomplete-input-suggestion-using-python-and-flask/
# https://www.geeksforgeeks.org/autocomplete-input-suggestion-using-python-and-flask/

max_rows=30       # Max number of movies returned
min_rating=5.0
min_votes=1000

df_ratings_titles=None
df_principals=None
df_names=None

#
# Logging
#
logging.basicConfig(
  filename='MoviesSearch.log', 
  # encoding='utf-8', 
  format='%(asctime)s %(levelname)s:%(message)s', 
  level=logging.DEBUG
)

app = Flask(__name__)

#
# Cross Origin to be able to call the API from another machine or port
#
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

# Use Flask server session to avoid a "Confirm Form Resubmission" pop-up:
# Redirect and pass form values from post to get method
app.config['SECRET_KEY'] = "your_secret_key" 
app.config['SESSION_TYPE'] = 'filesystem' 
app.config['SESSION_PERMANENT']= False
app.config.from_object(__name__)
Session(app)

@app.route('/', methods=['GET','POST'])
@cross_origin()
def slash():
  print(f"slash - got request: {request.method}")
  
  # The 'extract' button was pressed
  if 'extract' in request.form:
    logging.debug("request.form:")
    logging.debug(request.form)
    results=movies(request.form).to_dict('records')
    logging.debug("MoviesSearch - Extracted results:")
    logging.debug(results)
  
  # Download Option
  elif 'download' in request.form and 'results' in session:
  
    # Create a workbook
    wb = Workbook()
    
    # Assign the sheets
    ws = wb.active
    ws.title = "Movies"
    print(ws.title)
    
    # Get the data
    j=session['results']
    if j:
      # results=json.loads(session['results'])
      results=session['results']
    else:
      results=[]
    
    # Set the sheets
    set_sheet(results,ws)
    
    return Response(
      save_virtual_workbook(wb),
      headers={
        'Content-Disposition': 'attachment; filename=MoviesSearch.xlsx',
        'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
      }
    ) 
  # download
    
  # Redirect
  if request.method=='POST':
    logging.debug("MoviesSearch - branch: redirect")
    # for result in results:
    #   print(f"results 3 - k:{result['title']} {result['year']}")
    # session['results'] = json.dumps(results)
    session['results'] = results
    return redirect(url_for('slash'))

  # Render
  else:
    logging.debug("MoviesSearch - branch: render index.html")
    if 'results' in session:
      j=session['results']
      if j:
        # results=json.loads(session['results'])
        results=session['results']
      else:
        results=[]
    else:
      results=[]
    logging.debug("MoviesSearch - Rendered results:")
    logging.debug(results)

    return render_template("index.html",results=results)
# slash()

@app.route('/find_names', methods=['GET'])
@cross_origin()
def find_names():
    # print(f"find_names - got request: {request.method}")
    search = request.args.get('q')
    # logging.debug(f"find_names - search: {search}")
    df=df_names[df_names.primaryName.str.contains(search,na=False, case=False)]
    df=df.assign(display=df.primaryName+' ('+df.birthYear+')')
    return df.to_json(orient='records')
# find_names()

def movies(params):
    genre=""
    if "genre" in params.keys():
      genre=params['genre']
    nconst1=""
    if "name" in params.keys():
      nconst1=params['name']
    nconst2=""
    if "name2" in params.keys():
      nconst2=params['name2']
    if "rating" in params.keys() and params['rating'].isdigit():
      rating=float(d['rating'])
    else:
      rating=0
    title='title'
    if "votes" in params.keys() and params['votes'].isdigit():
      votes=int(params['votes'])
    else:
      votes=0
    if "yearEnd" in params.keys() and params['yearEnd'].isdigit():
      year_end=int(params['yearEnd'])
    else:
      year_end=99999
    if "yearStart" in params.keys() and params['yearStart'].isdigit():
      year_start=int(params['yearStart'])
    else:
      year_start=0
      
    logging.debug(f"movies - g:{genre} c1:{nconst1} c2:{nconst2} r:{rating} v:{votes} e:{year_end} s:{year_start}")

    df=find_titles(
            genre,
            nconst1,
            nconst2,
            rating,
            title,
            votes,
            year_end,
            year_start
    )
    # logging.debug("To json")
    # js=df.to_json(orient='records')
    # logging.debug(js)
    # # resp={'message': "Processed"}
    # return js
    # # return jsonify(resp)
    return df
# movies

def find_titles(genre,nconst1,nconst2,rating,title,votes,year_end,year_start):
    # Starting point cases
    # - No movie for name1 or Name2
    # - No movie for name1
    # - No movie for name2
    # - Movie(s) for name1 and name2
    
    logging.debug(f"find_titles: g:{genre} c1:{nconst1} c2:{nconst2} r:{rating} t:{title} v:{votes} ys:{year_start} ye:{year_end}")
    filtered=""
    df=df_ratings_titles
    logging.debug(f"find_titles - titles         : {df.shape}")
    if nconst1!="":
        df1=df_principals[df_principals.nconst==nconst1]
        logging.debug(f"find_titles - principals n1={nconst1}  : {df1.shape}")
        l1=df1['tconst'].unique().tolist()
        df=df[df.tconst.isin(l1)]
        logging.debug(f"find_titles - titles     n1={nconst1}  : {df.shape}")
    if nconst2!="":
        df2=df_principals[df_principals.nconst==nconst2]
        logging.debug(f"find_titles - principals n2={nconst2}  : {df2.shape}")
        l2=df2['tconst'].unique().tolist()
        df=df[df.tconst.isin(l2)]
        logging.debug(f"find_titles - titles     n2={nconst2}  : {df.shape}")
    logging.debug(f"find_titles - titles                   : {df.shape}")
    
    if genre!="":
        logging.debug(f"find_titles filter genre: {genre}")
        filtered+=" g:"+genre
        df=df.dropna(subset=['genres'])
        df=df[df.genres.str.contains(genre)]
    if rating > 0:
        filtered+=" r:"+str(rating)
        df=df[pd.to_numeric(df['averageRating'], errors='coerce').notnull()]
        df=df[df.averageRating > rating]
    # if title!="":
    #    filtered+=" t:"+title
    #    df=df.dropna(subset=['originalTitle','primaryTitle'])
    #    df=df[df.originalTitle.str.contains(title) or df.primaryTitle.str.contains(title)]
    if votes > 0:
        filtered+=" v:"+str(votes)
        df=df[pd.to_numeric(df['numVotes'], errors='coerce').notnull()]
        df=df[df.numVotes > votes]
    if year_end > 0:
        filtered+=" ye:"+str(year_end)
        df=df[pd.to_numeric(df['startYear'], errors='coerce').notnull()]
        df=df[df.startYear.astype(float).astype('Int64', errors='ignore') <= year_end]
    if year_start > 0:
        filtered+=" ys:"+str(year_start)
        df=df[pd.to_numeric(df['startYear'], errors='coerce').notnull()]
        df=df[df.startYear.astype(float).astype('Int64', errors='ignore') >= year_start]
    print(f"Filterd: {filtered}")

    print("Movies")
    print(df)
    
    # Casts
    l=df['tconst'].unique().tolist()
    df_casts=df_principals[df_principals['tconst'].isin(l)]
    print("Casts:")
    print(df_casts)
    df_casts_names = pd.merge(df_casts,df_names,on="nconst",how="left")
    print("Casts_names - joined:")
    print(df_casts_names)
    
    # Keep for director or searched actor
    df_casts_names=df_casts_names.dropna(subset=['category'])
    df_casts_names=df_casts_names[df_casts_names.category=='director']
    # df_casts=df_casts[df_casts.category=='director' or (nconst1!="" and df_casts.primaryName==nConst1) or (nconst2!="" and df_casts.primaryName==nConst2)]

    df_movies_casts_names = pd.merge(df_casts_names,df,on="tconst",how="left")
    print("Movies Casts_names - joined:")
    print(df_movies_casts_names)
    
    return df_movies_casts_names.head(max_rows)
# Find titles

def init():
    print("init()")
    logging.info('init()')
    names()
    principals()
    ratings_titles()
    global params
    # params=f"init(): Names: {df_names.size} Principals: {df_principals} Titles: {df_ratings_titles}"
    # logging.info(params)
# init

def list_genres():
    l=[]
    print("list_genres")
    g=df_ratings_titles['genres'].unique().tolist()
    for genres in g:
      if isinstance(genres, str):
        tokens=genres.split(',')
        for t in tokens:
          # print(t)
          if not t in l:
            l.append(t)
      # print(genres)
    return l
# list_genres

def names():
    global df_names
    print("Load Names...")
    df_names=pd.read_csv("names.min")
    # display(df_names)
# names

def principals():
    global df_principals
    print("Load Principals...")
    df_principals=pd.read_csv("principals.min")
    # display(df_principals)
# principals

def ratings_titles():
    global df_ratings_titles
    print("Load Ratings and Titles..")
    file="ratings_titles.min"
    df_ratings_titles = pd.read_csv(file)
    # display(df_ratings_titles)
# ratings_titles
    
def set_sheet(data,ws):

    c=1
    r=1
      
    # Headers
    ws[get_column_letter(c)+str(r)]='Movie Title'
    ws[get_column_letter(c+1)+str(r)]='Year'
    ws[get_column_letter(c+2)+str(r)]='Director'
    ws[get_column_letter(c+3)+str(r)]='Rating'
      
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
      
    # Data
    logging.debug("set_sheet - data")
    logging.debug(data)
    for row in data:
      logging.debug("set_sheet - row")
      logging.debug(row)
      r+=1
      ws[get_column_letter(c)+str(r)]=row['originalTitle']
      ws[get_column_letter(c+1)+str(r)]=row['startYear']
      ws[get_column_letter(c+2)+str(r)]=row['primaryName']
      ws[get_column_letter(c+3)+str(r)]=row['averageRating']
        
# set_sheet

if __name__ == '__main__':

    init()
    
    # Find Titles
    genre=""
    name1="" # "nm0001191"
    name2="" # "nm0001705"
    rating=5.5
    title=""
    votes=1000
    year_start=2021
    year_end=2022
    # df = find_titles(genre,name1,name2,rating,title,votes,year_end,year_start)
    # df.to_csv("movies_casts_names.csv")

    # Find name
    # search="Marie-Josée Croze"
    # search="Marie"
    # df = find_names(search,10)
    # df.to_csv("find_name.csv")
    
    # To open port:
    # Type "Windows Firewall" in search
    # Go to Windows Defender Firewall
    # > Advanced Settings > Inbound rules
    # > New Rule > Port > Next
    app.run(debug=True,host='0.0.0.0',port=5004)
